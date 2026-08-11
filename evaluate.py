#!/usr/bin/env python
"""
RVC Voice Model Evaluator — Main Entry Point.

Evaluates multiple RVC voice models against an original reference audio
using ECAPA-TDNN speaker similarity, acoustic analysis, F0/prosody,
stability, and artifact detection.

Usage:
    python evaluate.py
    python evaluate.py --original ./original.wav --input ./models
    python evaluate.py --config ./config.yaml
    python evaluate.py --device cuda
    python evaluate.py --device cpu
    python evaluate.py --self-test
    python evaluate.py --compare 24 25
    python evaluate.py --trend-dir ./results
    python evaluate.py --help
"""

import argparse
import csv
import json
import logging
import platform
import random
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

import numpy as np
import yaml
import torch

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from audio import (
    load_audio, prepare_analysis_audio, prepare_raw_audio,
    split_into_windows, trim_to_common_duration,
    load_model_files,
)
from vad import filter_valid_windows, get_overall_speech_ratio
from ecapa import ECAPAModel, compute_ecapa_scores
from acoustic import compute_acoustic_scores, compute_per_window_acoustic_scores
from f0 import compute_f0_scores, compute_per_window_f0_scores
from stability import compute_stability, compute_window_scores
from artifacts import compute_artifact_metrics, detect_absolute_warnings
from scoring import compute_final_scores
from visualization import plot_ranking, plot_radar_top5
from report import generate_html_report
from comparison import ModelComparator
from trend_analysis import TrendAnalyzer

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("evaluate")


def load_config(config_path: str) -> Dict[str, Any]:
    """Load YAML configuration."""
    path = Path(config_path)
    if not path.exists():
        logger.error("Config file not found: %s", config_path)
        sys.exit(1)

    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    logger.info("Configuration loaded from %s", config_path)
    return cfg


def set_seed(seed: int, deterministic: bool = True):
    """Set random seeds for reproducibility."""
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
    if deterministic:
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False
    logger.info("Random seed set to %d (deterministic=%s)", seed, deterministic)


def get_runtime_info(config: Dict[str, Any]) -> Dict[str, Any]:
    """Collect runtime environment information."""
    info = {
        "python_version": platform.python_version(),
        "pytorch_version": torch.__version__,
        "cuda_version": torch.version.cuda if torch.cuda.is_available() else "N/A",
        "cuda_available": torch.cuda.is_available(),
        "device": config.get("runtime", {}).get("device", "auto"),
    }

    try:
        import librosa
        info["librosa_version"] = librosa.__version__
    except ImportError:
        info["librosa_version"] = "N/A"

    try:
        import soundfile
        info["soundfile_version"] = soundfile.__version__
    except ImportError:
        info["soundfile_version"] = "N/A"

    try:
        import scipy
        info["scipy_version"] = scipy.__version__
    except ImportError:
        info["scipy_version"] = "N/A"

    try:
        import torchaudio
        info["torchaudio_version"] = torchaudio.__version__
    except ImportError:
        info["torchaudio_version"] = "N/A"

    return info


def self_test(config: Dict[str, Any]) -> None:
    """Run self-baseline test: compare original against itself.

    Splits original.wav into two halves and evaluates speaker similarity
    and acoustic distances to establish baseline values.
    """
    logger.info("=" * 60)
    logger.info("SELF-BASELINE TEST: Original vs Original")
    logger.info("=" * 60)

    original_path = config.get("original_audio", "./original.wav")
    if not Path(original_path).exists():
        logger.error("Original audio not found for self-test: %s", original_path)
        return

    sr = config.get("audio", {}).get("sample_rate", 16000)
    window_sec = config.get("window", {}).get("seconds", 3.0)
    min_windows = config.get("self_baseline", {}).get("min_windows", 3)

    # Load and split original
    audio, actual_sr = load_audio(original_path, target_sr=sr, mono=True)
    logger.info("Loaded: %s, duration=%.2fs, sr=%d", original_path, len(audio)/sr, actual_sr)

    mid = len(audio) // 2
    audio_a = audio[:mid]
    audio_b = audio[mid:]

    # Split into windows
    windows_a = split_into_windows(
        audio_a, sr, window_seconds=window_sec,
        drop_incomplete=True, min_valid_seconds=2.0,
    )
    windows_b = split_into_windows(
        audio_b, sr, window_seconds=window_sec,
        drop_incomplete=True, min_valid_seconds=2.0,
    )

    # VAD
    windows_a = filter_valid_windows(windows_a, sr, config.get("vad", {}).get("min_speech_ratio", 0.05))
    windows_b = filter_valid_windows(windows_b, sr, config.get("vad", {}).get("min_speech_ratio", 0.05))

    valid_a = [w for w in windows_a if w.get("valid", False)]
    valid_b = [w for w in windows_b if w.get("valid", False)]

    if len(valid_a) < min_windows or len(valid_b) < min_windows:
        logger.error("Not enough valid windows for self-test (need %d): A=%d, B=%d",
                     min_windows, len(valid_a), len(valid_b))
        return

    n = min(len(valid_a), len(valid_b))
    valid_a = valid_a[:n]
    valid_b = valid_b[:n]

    # ECAPA
    ecapa_cfg = config.get("ecapa", {})
    ecapa_dir = ecapa_cfg.get("model_dir", "./ecapa_model")
    ecapa = ECAPAModel(ecapa_dir, config.get("runtime", {}).get("device", "auto"))

    ecapa_sims = []
    mcd_vals = []
    mel_vals = []
    mfcc_vals = []

    if ecapa.check_model_exists():
        try:
            ecapa.load()
        except Exception as e:
            logger.error("ECAPA model load failed in self-test: %s", e)
        else:
            for wa, wb in zip(valid_a, valid_b):
                sim = _window_ecapa(ecapa, wa["audio"], wb["audio"])
                if sim is not None:
                    ecapa_sims.append(sim)
    else:
        logger.warning("ECAPA model not found — skipping ECAPA self-test")

    # Acoustic
    feat_cfg = config.get("features", {})
    from acoustic import compute_mcd, compute_mel_distance, compute_mfcc_distance
    logger.info("Computing per-window acoustic self-distances...")
    for wa, wb in zip(valid_a, valid_b):
        try:
            mcd_vals.append(compute_mcd(
                wa["audio"], wb["audio"], sr,
                n_mfcc=24,
                n_fft=feat_cfg.get("n_fft", 1024),
                hop_length=feat_cfg.get("hop_length", 256),
                win_length=feat_cfg.get("win_length", 1024),
            ))
        except Exception as e:
            logger.debug("MCD window computation skipped: %s", e)
        try:
            mel_vals.append(compute_mel_distance(
                wa["audio"], wb["audio"], sr,
                n_fft=feat_cfg.get("n_fft", 1024),
                hop_length=feat_cfg.get("hop_length", 256),
                win_length=feat_cfg.get("win_length", 1024),
            ))
        except Exception as e:
            logger.debug("Mel distance window computation skipped: %s", e)
        try:
            mfcc_vals.append(compute_mfcc_distance(
                wa["audio"], wb["audio"], sr,
                n_fft=feat_cfg.get("n_fft", 1024),
                hop_length=feat_cfg.get("hop_length", 256),
                win_length=feat_cfg.get("win_length", 1024),
            ))
        except Exception as e:
            logger.debug("MFCC distance window computation skipped: %s", e)

    from f0 import extract_f0, compute_f0_correlation
    logger.info("Computing per-window F0 self-correlations...")
    f0_sims = []
    for wa, wb in zip(valid_a, valid_b):
        try:
            f0a, voa = extract_f0(wa["audio"], sr)
            f0b, vob = extract_f0(wb["audio"], sr)
            min_len = min(len(f0a), len(f0b))
            merged_mask = voa[:min_len] & vob[:min_len]
            corr = compute_f0_correlation(f0a[:min_len], f0b[:min_len], merged_mask)
            if corr is not None:
                f0_sims.append(corr)
        except Exception as e:
            logger.debug("F0 window computation skipped: %s", e)

    # Output
    results = {
        "test": "self_baseline",
        "original": str(original_path),
        "num_window_pairs": n,
        "ecapa_self_similarity": float(np.mean(ecapa_sims)) if ecapa_sims else None,
        "ecapa_self_similarity_std": float(np.std(ecapa_sims)) if ecapa_sims else None,
        "ecapa_num_windows": len(ecapa_sims),
        "mcd_self_distance": float(np.mean(mcd_vals)) if mcd_vals else None,
        "mcd_self_distance_std": float(np.std(mcd_vals)) if mcd_vals else None,
        "mel_self_distance": float(np.mean(mel_vals)) if mel_vals else None,
        "mel_self_distance_std": float(np.std(mel_vals)) if mel_vals else None,
        "mfcc_self_distance": float(np.mean(mfcc_vals)) if mfcc_vals else None,
        "mfcc_self_distance_std": float(np.std(mfcc_vals)) if mfcc_vals else None,
        "f0_self_similarity": float(np.mean(f0_sims)) if f0_sims else None,
        "f0_self_similarity_std": float(np.std(f0_sims)) if f0_sims else None,
    }

    import json
    import json
    results_dir = Path(config.get("results_dir", "./results")) / "self_baseline"
    results_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(results_dir / "self_baseline.json")
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2, default=str)
    logger.info("Self-baseline saved to %s", output_path)
    logger.info("ECAPA self-similarity: %s",
                f"{results['ecapa_self_similarity']:.4f}" if results['ecapa_self_similarity'] else "N/A")
    logger.info("MCD self-distance: %s",
                f"{results['mcd_self_distance']:.4f}" if results['mcd_self_distance'] else "N/A")
    logger.info("Mel self-distance: %s",
                f"{results['mel_self_distance']:.4f}" if results['mel_self_distance'] else "N/A")
    logger.info("MFCC self-distance: %s",
                f"{results['mfcc_self_distance']:.4f}" if results['mfcc_self_distance'] else "N/A")
    logger.info("F0 self-similarity: %s",
                f"{results['f0_self_similarity']:.4f}" if results['f0_self_similarity'] else "N/A")


def _window_ecapa(ecapa_model, audio_a: np.ndarray, audio_b: np.ndarray) -> float:
    """Compute ECAPA similarity for two audio windows."""
    try:
        embs = ecapa_model.encode_batch([audio_a, audio_b])
        sim = ecapa_model.compute_similarity(embs[0], embs[1])
        return sim
    except Exception:
        return None


def evaluate_models(config: Dict[str, Any], original_path: str, models_dir: str, results_dir: str) -> Path:
    """Main evaluation pipeline."""
    logger.info("=" * 60)
    logger.info("RVC VOICE MODEL EVALUATION")
    logger.info("=" * 60)

    # ---- Setup ----
    sr = config.get("audio", {}).get("sample_rate", 16000)
    normalize_analysis = config.get("audio", {}).get("normalize_for_analysis", True)
    window_sec = config.get("window", {}).get("seconds", 3.0)
    min_speech = config.get("vad", {}).get("min_speech_ratio", 0.05)

    # ---- Load original ----
    if not Path(original_path).exists():
        logger.error("Original audio not found: %s", original_path)
        sys.exit(1)

    orig_analysis, _ = prepare_analysis_audio(original_path, sr, normalize=normalize_analysis)
    orig_raw, _ = prepare_raw_audio(original_path, sr)
    logger.info("Original loaded: %s, analysis_dur=%.2fs", original_path, len(orig_analysis)/sr)

    # Compute artifact metrics for original audio (used as reference for delta scoring)
    orig_artifacts = compute_artifact_metrics(orig_raw, sr)
    logger.info("Original artifacts: SF=%.4f, HF=%.4f, Clip=%.4f",
                orig_artifacts.get("spectral_flatness", -1),
                orig_artifacts.get("high_frequency_ratio", -1),
                orig_artifacts.get("clipping_ratio", -1))

    # ---- Scan models ----
    model_files = load_model_files(models_dir)
    if not model_files:
        logger.error("No model WAV files found in %s", models_dir)
        sys.exit(1)

    logger.info("Found %d model(s) to evaluate", len(model_files))

    # ---- ECAPA model ----
    ecapa_cfg = config.get("ecapa", {})
    ecapa_dir = ecapa_cfg.get("model_dir", "./ecapa_model")
    ecapa = ECAPAModel(ecapa_dir, config.get("runtime", {}).get("device", "auto"))

    ecapa_available = ecapa.check_model_exists()
    if ecapa_available:
        try:
            ecapa.load()
        except Exception as e:
            logger.error("ECAPA model load failed: %s — ECAPA scores will be NaN", e)
            ecapa_available = False
    else:
        logger.warning("ECAPA model not available — ECAPA scores will be NaN")

    # ---- Process each model ----
    all_results = []

    for mf in model_files:
        model_name = mf.stem
        logger.info("--- Processing: %s ---", model_name)

        result = {"model_name": model_name, "model_path": str(mf)}

        try:
            # Analysis pipeline
            model_analysis, _ = prepare_analysis_audio(str(mf), sr, normalize=normalize_analysis)
            model_raw, _ = prepare_raw_audio(str(mf), sr)

            # Trim to common duration
            orig_trim, model_trim = trim_to_common_duration(orig_analysis, model_analysis, sr)

            usable_dur = len(orig_trim) / sr
            result["duration"] = usable_dur
            logger.info("  Usable duration: %.2fs", usable_dur)

            # Split into windows
            orig_windows = split_into_windows(
                orig_trim, sr, window_seconds=window_sec,
                drop_incomplete=True, min_valid_seconds=config.get("window", {}).get("min_valid_window_seconds", 2.0),
            )
            model_windows = split_into_windows(
                model_trim, sr, window_seconds=window_sec,
                drop_incomplete=True, min_valid_seconds=config.get("window", {}).get("min_valid_window_seconds", 2.0),
            )

            n_windows = min(len(orig_windows), len(model_windows))
            orig_windows = orig_windows[:n_windows]
            model_windows = model_windows[:n_windows]
            result["num_windows"] = n_windows
            logger.info("  Windows: %d", n_windows)

            if n_windows == 0:
                logger.warning("  No valid windows for %s — skipping", model_name)
                result["status"] = "SKIPPED: No windows"
                all_results.append(result)
                continue

            # ---- VAD ----
            orig_windows = filter_valid_windows(orig_windows, sr, min_speech)
            model_windows = filter_valid_windows(model_windows, sr, min_speech)

            speech_ratio = get_overall_speech_ratio(model_windows)
            result["speech_ratio"] = speech_ratio
            logger.info("  Speech ratio: %.2f", speech_ratio)

            # ---- ECAPA ----
            if ecapa_available:
                ecapa_result = compute_ecapa_scores(
                    orig_windows, model_windows, ecapa,
                    mean_weight=ecapa_cfg.get("mean_weight", 0.75),
                    min_weight=ecapa_cfg.get("min_weight", 0.15),
                )
                result["ecapa"] = ecapa_result
                logger.info("  ECAPA score: %.4f", ecapa_result.get("ecapa_score", 0) or 0)
            else:
                result["ecapa"] = {"status": "UNAVAILABLE", "ecapa_score": None}

            # ---- Acoustic ----
            try:
                acoustic_result = compute_acoustic_scores(orig_trim, model_trim, sr, config)
                result["acoustic"] = acoustic_result
                logger.info("  MCD=%.4f, Mel=%.4f, MFCC=%.4f",
                           acoustic_result.get("mcd_distance", -1) or -1,
                           acoustic_result.get("mel_distance", -1) or -1,
                           acoustic_result.get("mfcc_distance", -1) or -1)
            except Exception as e:
                logger.warning("  Acoustic failed: %s", e)
                result["acoustic"] = {"status": f"FAILED: {e}", "mcd_distance": None, "mel_distance": None, "mfcc_distance": None}

            # ---- F0 ----
            try:
                f0_result = compute_f0_scores(orig_trim, model_trim, sr, config)
                result["f0"] = f0_result
                logger.info("  F0 corr=%.4f, nRMSE=%.4f",
                           f0_result.get("f0_correlation", -1) or -1,
                           f0_result.get("f0_rmse", -1) or -1)
            except Exception as e:
                logger.warning("  F0 failed: %s", e)
                result["f0"] = {"status": f"FAILED: {e}", "f0_correlation": None, "f0_rmse": None, "f0_dtw": None}

            # ---- Stability (real window-level composite scores) ----
            ecapa_per_win = result.get("ecapa", {}).get("per_window_similarities", [])

            # Use same valid indices as ECAPA
            if ecapa_per_win:
                valid_indices = [
                    i for i, (ow, mw) in enumerate(zip(orig_windows, model_windows))
                    if ow.get("valid", False) and mw.get("valid", False)
                ]
                n_valid = len(valid_indices)

                if n_valid > 0:
                    # Align ecapa_per_win with valid_indices (they should match in order)
                    n_ecapa = len(ecapa_per_win)
                    use_n = min(n_valid, n_ecapa)
                    orig_win_audio = [orig_windows[valid_indices[i]]["audio"] for i in range(use_n)]
                    model_win_audio = [model_windows[valid_indices[i]]["audio"] for i in range(use_n)]
                    ecapa_for_stability = ecapa_per_win[:use_n]

                    # Per-window acoustic scores
                    try:
                        win_acoustic = compute_per_window_acoustic_scores(
                            orig_win_audio, model_win_audio, sr, config,
                        )
                    except Exception as e:
                        logger.warning("  Per-window acoustic failed: %s", e)
                        win_acoustic = [np.nan] * use_n

                    # Per-window F0 scores
                    try:
                        win_f0 = compute_per_window_f0_scores(
                            orig_win_audio, model_win_audio, sr, config,
                        )
                    except Exception as e:
                        logger.warning("  Per-window F0 failed: %s", e)
                        win_f0 = [np.nan] * use_n

                    # Composite window scores (ECAPA + Acoustic + F0)
                    st_cfg = config.get("final_score", {})
                    window_scores = compute_window_scores(
                        ecapa_for_stability, win_acoustic, win_f0,
                        ecapa_weight=st_cfg.get("ecapa_weight", 0.40),
                        acoustic_weight=st_cfg.get("acoustic_weight", 0.30),
                        f0_weight=st_cfg.get("f0_weight", 0.15),
                    )
                    stability_result = compute_stability(window_scores)
                    logger.info("  Stability: mean=%.4f, std=%.4f (n_valid=%d)",
                               stability_result.get("window_score_mean", -1) or -1,
                               stability_result.get("window_score_std", -1) or -1,
                               stability_result.get("n_valid_windows", 0))
                else:
                    stability_result = {
                        "window_score_mean": None, "window_score_std": None,
                        "window_score_min": None, "window_score_max": None,
                        "window_score_range": None,
                        "n_windows": 0, "n_valid_windows": 0,
                        "status": "FAILED: No valid windows",
                    }
            else:
                stability_result = {
                    "window_score_mean": None, "window_score_std": None,
                    "window_score_min": None, "window_score_max": None,
                    "window_score_range": None,
                    "n_windows": 0, "n_valid_windows": 0,
                    "status": "FAILED: No ECAPA windows",
                }
            result["stability"] = stability_result

            # ---- Artifacts (on FULL raw audio, NOT trimmed) ----
            try:
                artifact_result = compute_artifact_metrics(model_raw, sr)
                # Store original artifacts for delta comparison
                artifact_result["orig_spectral_flatness"] = orig_artifacts.get("spectral_flatness")
                artifact_result["orig_high_frequency_ratio"] = orig_artifacts.get("high_frequency_ratio")
                artifact_result["orig_clipping_ratio"] = orig_artifacts.get("clipping_ratio")
                artifact_result["orig_peak"] = orig_artifacts.get("peak")
                artifact_result["orig_rms"] = orig_artifacts.get("rms")
                result["artifacts"] = artifact_result
                logger.info("  Clip=%.4f, HF=%.4f, SF=%.4f",
                           artifact_result.get("clipping_ratio", -1),
                           artifact_result.get("high_frequency_ratio", -1),
                           artifact_result.get("spectral_flatness", -1))
            except Exception as e:
                logger.warning("  Artifact detection failed: %s", e)
                result["artifacts"] = {"status": f"FAILED: {e}"}

            # ---- Absolute warnings ----
            artifacts = result.get("artifacts", {})
            clipping = artifacts.get("clipping_ratio")
            hf_ratio = artifacts.get("high_frequency_ratio")
            flatness = artifacts.get("spectral_flatness")
            if any(v is None for v in [clipping, hf_ratio, flatness]):
                result["warnings"] = ["ARTIFACT_METRICS_UNAVAILABLE"]
            else:
                result["warnings"] = detect_absolute_warnings(
                    clipping, hf_ratio, flatness, speech_ratio,
                )
            if result["warnings"]:
                logger.warning("  Warnings: %s", result["warnings"])

            result["status"] = "OK"

        except Exception as e:
            logger.error("  FAILED for %s: %s", model_name, e, exc_info=True)
            result["status"] = f"FAILED: {e}"

        all_results.append(result)

    # ---- Batch Scoring ----
    logger.info("=" * 60)
    logger.info("COMPUTING BATCH SCORES (Min-Max Normalization)")
    logger.info("=" * 60)

    all_results = compute_final_scores(all_results, config)

    # Print ranking
    tech_sorted = sorted(all_results, key=lambda r: r["scores"].get("technical_rank", 999))
    logger.info("FINAL RANKING:")
    for r in tech_sorted:
        s = r["scores"]
        fs = s["final_score"]
        fs_str = f"{fs:.2f}" if not np.isnan(fs) else "N/A"
        warnings = s.get("warnings", [])
        w_str = f" [WARN] {', '.join(warnings)}" if warnings else ""
        logger.info("  #%d Tech / #%d Rec | %s | %s%s",
                   s.get("technical_rank", "?"),
                   s.get("recommended_rank", "?"),
                   r["model_name"], fs_str, w_str)

    # ---- Generate Outputs ----
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(results_dir) / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)
    logger.info("Output directory: %s", output_dir)

    # Save outputs
    _save_outputs(all_results, config, output_dir, timestamp)

    logger.info("=" * 60)
    logger.info("EVALUATION COMPLETE")
    logger.info("Results: %s", output_dir)
    logger.info("=" * 60)
    return output_dir


# ---- Post-evaluation tools ----


def run_comparison(result_dir: Path, model_a: str, model_b: str) -> None:
    """Run A/B comparison on the just-completed evaluation results."""
    scores_file = result_dir / "scores.json"
    if not scores_file.exists():
        logger.error("scores.json not found in %s — cannot compare", result_dir)
        return

    comparator = ModelComparator()
    try:
        ma = comparator.load_model(str(scores_file), model_a)
        mb = comparator.load_model(str(scores_file), model_b)
    except ValueError as e:
        logger.error("Comparison failed: %s", e)
        return

    result = comparator.compare(ma, mb)

    # Print summary
    print(f"\n{'='*50}")
    print(f"A/B Comparison: {model_a}  vs  {model_b}")
    print(f"{'='*50}")
    for metric in ModelComparator.METRICS:
        diff = result["difference"][metric]
        winner = result["winner"][metric]
        arrow = " →" if diff > 0 else ("← " if diff < 0 else " ≈")
        print(f"  {metric:<12} {arrow} {diff:+.3f}  ({winner})")

    print(f"\n--- Analysis ---")
    for line in result["analysis"]:
        print(f"  · {line}")

    overall = result["difference"]["score"]
    if overall > 0:
        print(f"\nConclusion: Model B ({result['model_b']}) outperforms Model A ({result['model_a']})")
    elif overall < 0:
        print(f"\nConclusion: Model A ({result['model_a']}) outperforms Model B ({result['model_b']})")
    else:
        print(f"\nConclusion: Models perform equally")

    # Save
    cmp_path = result_dir / f"comparison_{model_a}_vs_{model_b}.json"
    comparator.save_report(result, str(cmp_path))
    logger.info("Comparison saved: %s", cmp_path)


def run_trend_analysis(folder: str) -> None:
    """Run overfit trend analysis on a folder of evaluation results."""
    from trend_analysis import TrendAnalyzer

    analyzer = TrendAnalyzer()
    reports = analyzer.load_reports(folder)

    if not reports:
        logger.error("No model reports found in %s", folder)
        return

    print(f"\nLoaded {len(reports)} checkpoint(s) from {folder}")

    trend = analyzer.analyze_trend(reports)
    print(f"\n=== Score Trend ===")
    for t in trend:
        bar = "|" * max(1, int(t["score"] / 5))
        print(f"  {str(t['epoch']):>6}  {t['score']:>6.2f}  {bar}")

    result = analyzer.explain(reports)
    print(f"\n=== Summary ===")
    for line in result["summary"]:
        print(f"  {line}")

    out_path = Path(folder) / "trend_analysis.json"
    result["trend"] = trend
    analyzer.save(result, str(out_path))
    logger.info("Trend analysis saved: %s", out_path)


def _save_outputs(
    all_results: List[Dict[str, Any]],
    config: Dict[str, Any],
    output_dir: Path,
    timestamp: str,
) -> None:
    """Save all output files."""
    out_cfg = config.get("output", {})
    top_n = out_cfg.get("top_n", 5)

    # Sort by technical rank
    sorted_results = sorted(all_results, key=lambda r: r["scores"].get("technical_rank", 999))

    # --- CSV ---
    if out_cfg.get("save_csv", True):
        csv_path = output_dir / "ranking.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "rank", "model", "final_score", "technical_rank", "recommended_rank",
                "ecapa_mean", "ecapa_std", "ecapa_min", "ecapa_max", "ecapa_score",
                "mcd_distance", "mcd_score", "mel_distance", "mel_score",
                "mfcc_distance", "mfcc_score", "acoustic_score",
                "f0_correlation", "f0_rmse", "f0_dtw", "f0_score",
                "window_score_mean", "window_score_std", "stability_score",
                "clipping_ratio", "peak", "rms", "spectral_flatness",
                "high_frequency_ratio", "artifact_score",
                "speech_ratio", "duration", "num_windows",
                "absolute_score", "absolute_level",
                "missing_metrics", "warnings", "status",
            ])
            for i, r in enumerate(sorted_results, 1):
                s = r["scores"]
                writer.writerow([
                    i,
                    r["model_name"],
                    s.get("final_score"),
                    s.get("technical_rank"),
                    s.get("recommended_rank"),
                    s.get("ecapa_mean"), s.get("ecapa_std"),
                    s.get("ecapa_min"), s.get("ecapa_max"), s.get("ecapa_score"),
                    s.get("mcd_distance"), s.get("mcd_score"),
                    s.get("mel_distance"), s.get("mel_score"),
                    s.get("mfcc_distance"), s.get("mfcc_score"),
                    s.get("acoustic_score"),
                    s.get("f0_correlation"), s.get("f0_rmse"), s.get("f0_dtw"),
                    s.get("f0_score"),
                    s.get("window_score_mean"), s.get("window_score_std"),
                    s.get("stability_score"),
                    s.get("clipping_ratio"), s.get("peak"), s.get("rms"),
                    s.get("spectral_flatness"), s.get("high_frequency_ratio"),
                    s.get("artifact_score"),
                    s.get("speech_ratio"), s.get("duration"), s.get("num_windows"),
                    (s.get("absolute_score") or {}).get("score", ""),
                    (s.get("absolute_score") or {}).get("overall_level", ""),
                    ";".join(s.get("missing_metrics", [])),
                    ";".join(s.get("warnings", [])),
                    s.get("status"),
                ])
        logger.info("CSV saved: %s", csv_path)

    # --- Detailed CSV ---
    if out_cfg.get("save_detailed_csv", True):
        det_path = output_dir / "detailed_scores.csv"
        with open(det_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "model", "final_score", "ecapa_mean", "ecapa_std", "ecapa_min", "ecapa_max",
                "mcd_distance", "mel_distance", "mfcc_distance",
                "f0_correlation", "f0_rmse", "f0_dtw",
                "clipping_ratio", "rms", "spectral_flatness", "high_frequency_ratio",
                "speech_ratio", "duration",
            ])
            for r in sorted_results:
                s = r["scores"]
                writer.writerow([
                    r["model_name"], s.get("final_score"),
                    s.get("ecapa_mean"), s.get("ecapa_std"),
                    s.get("ecapa_min"), s.get("ecapa_max"),
                    s.get("mcd_distance"), s.get("mel_distance"), s.get("mfcc_distance"),
                    s.get("f0_correlation"), s.get("f0_rmse"), s.get("f0_dtw"),
                    s.get("clipping_ratio"), s.get("rms"),
                    s.get("spectral_flatness"), s.get("high_frequency_ratio"),
                    s.get("speech_ratio"), s.get("duration"),
                ])

    # --- Excel ---
    if out_cfg.get("save_excel", True):
        _save_excel(all_results, output_dir / "ranking.xlsx", config)

    # --- JSON ---
    if out_cfg.get("save_json", True):
        json_path = output_dir / "scores.json"
        json_data = _prepare_json_data(all_results, config, timestamp)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
        logger.info("JSON saved: %s", json_path)

    # --- Charts ---
    ranking_png = str(output_dir / "ranking.png")
    radar_png = str(output_dir / "radar_top5.png")

    if out_cfg.get("save_ranking_plot", True):
        try:
            plot_ranking(all_results, ranking_png)
        except Exception as e:
            logger.warning("Ranking plot failed: %s", e)
            ranking_png = ""

    if out_cfg.get("save_radar_plot", True):
        try:
            plot_radar_top5(all_results, radar_png, top_n=top_n)
        except Exception as e:
            logger.warning("Radar plot failed: %s", e)
            radar_png = ""

    # --- HTML Report ---
    if out_cfg.get("save_html", True):
        runtime_info = get_runtime_info(config)
        try:
            generate_html_report(all_results, str(output_dir / "report.html"),
                                config, runtime_info, ranking_png, radar_png)
        except Exception as e:
            logger.warning("HTML report failed: %s", e)

    # --- Top 5 Audio ---
    if out_cfg.get("save_top5_audio", True):
        try:
            _save_top5_audio(all_results, output_dir, top_n)
        except Exception as e:
            logger.warning("Top 5 audio copy failed: %s", e)

    # --- Comparison Audio ---
    if out_cfg.get("save_comparison_audio", True):
        try:
            _save_comparison_audio(all_results, config, output_dir)
        except Exception as e:
            logger.warning("Comparison audio failed: %s", e)


def _save_excel(results: List[Dict[str, Any]], path: Path, config: Dict[str, Any]) -> None:
    """Save Excel file with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel output")
        return

    wb = openpyxl.Workbook()

    # Sheet 1: Ranking
    ws = wb.active
    ws.title = "Ranking"
    headers = ["Rank", "Model", "Final Score", "Tech Rank", "Rec Rank",
               "ECAPA", "Acoustic", "F0", "Stability", "Artifact",
               "Abs Score", "Abs Level", "Warnings"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    sorted_r = sorted(results, key=lambda r: r["scores"].get("technical_rank", 999))
    for i, r in enumerate(sorted_r, 1):
        s = r["scores"]
        ws.append([
            i, r["model_name"], s.get("final_score"),
            s.get("technical_rank"), s.get("recommended_rank"),
            s.get("ecapa_score"), s.get("acoustic_score"),
            s.get("f0_score"), s.get("stability_score"),
            s.get("artifact_score"),
            (s.get("absolute_score") or {}).get("score", ""),
            (s.get("absolute_score") or {}).get("overall_level", ""),
            "; ".join(s.get("warnings", [])),
        ])

    # Sheet 2: Detailed Scores
    ws2 = wb.create_sheet("Detailed Scores")
    det_headers = ["Model", "Final", "ECAPA Mean", "MCD", "Mel", "MFCC",
                   "F0 Corr", "F0 RMSE", "Clip", "HF Ratio", "Speech"]
    ws2.append(det_headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in sorted_r:
        s = r["scores"]
        ws2.append([
            r["model_name"], s.get("final_score"),
            s.get("ecapa_mean"), s.get("mcd_distance"),
            s.get("mel_distance"), s.get("mfcc_distance"),
            s.get("f0_correlation"), s.get("f0_rmse"),
            s.get("clipping_ratio"), s.get("high_frequency_ratio"),
            s.get("speech_ratio"),
        ])

    # Sheet 3: Configuration (config.yaml + environment info)
    ws3 = wb.create_sheet("Configuration")
    ws3.append(["Section", "Parameter", "Value"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Config parameters
    for k, v in _flatten_config(config).items():
        parts = k.split(".", 1)
        section = parts[0] if len(parts) > 1 else ""
        param = parts[1] if len(parts) > 1 else k
        ws3.append([section, param, str(v)])

    # Environment / runtime info
    runtime_info = get_runtime_info(config)
    ws3.append(["", "", ""])
    ws3.append(["Environment", "", ""])
    ws3.append(["", "python_version", runtime_info.get("python_version", "N/A")])
    ws3.append(["", "pytorch_version", runtime_info.get("pytorch_version", "N/A")])
    ws3.append(["", "torchaudio_version", runtime_info.get("torchaudio_version", "N/A")])
    ws3.append(["", "librosa_version", runtime_info.get("librosa_version", "N/A")])
    ws3.append(["", "scipy_version", runtime_info.get("scipy_version", "N/A")])
    ws3.append(["", "soundfile_version", runtime_info.get("soundfile_version", "N/A")])
    ws3.append(["", "cuda_version", runtime_info.get("cuda_version", "N/A")])
    ws3.append(["", "cuda_available", str(runtime_info.get("cuda_available", False))])
    ws3.append(["", "device", runtime_info.get("device", "auto")])

    # ECAPA model info
    ecapa_cfg = config.get("ecapa", {})
    ecapa_dir = ecapa_cfg.get("model_dir", "./ecapa_model")
    ws3.append(["ECAPA_Model", "", ""])
    ws3.append(["", "model_dir", ecapa_dir])
    ecapa_path = Path(ecapa_dir)
    if ecapa_path.exists():
        model_files = [f.name for f in ecapa_path.iterdir() if f.is_file()]
        ws3.append(["", "model_files", "; ".join(model_files) if model_files else "empty"])
    else:
        ws3.append(["", "model_files", "directory not found"])

    # Sheet 4: Human Evaluation
    ws4 = wb.create_sheet("Human Evaluation")
    ws4.append(["Model", "Technical Rank", "Recommended Rank", "Human Rating", "Human Notes"])
    for r in sorted_r:
        s = r["scores"]
        ws4.append([r["model_name"], s.get("technical_rank"), s.get("recommended_rank"), "", ""])

    wb.save(path)
    logger.info("Excel saved: %s", path)


def _flatten_config(cfg: Dict, prefix: str = "") -> Dict[str, Any]:
    """Flatten nested config dict."""
    result = {}
    for k, v in cfg.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            result.update(_flatten_config(v, key))
        else:
            result[key] = v
    return result


def _prepare_json_data(results: List[Dict[str, Any]], config: Dict[str, Any], timestamp: str) -> Dict:
    """Prepare JSON-serializable output data."""
    sorted_r = sorted(results, key=lambda r: r["scores"].get("technical_rank", 999))
    models_data = []
    for r in sorted_r:
        s = r["scores"]
        entry = {
            "model": r["model_name"],
            "path": r["model_path"],
            "final_score": s.get("final_score"),
            "technical_rank": s.get("technical_rank"),
            "recommended_rank": s.get("recommended_rank"),
            "sub_scores": {
                "ecapa": s.get("ecapa_score"),
                "acoustic": s.get("acoustic_score"),
                "f0": s.get("f0_score"),
                "stability": s.get("stability_score"),
                "artifact": s.get("artifact_score"),
            },
            "raw_metrics": {
                "ecapa_mean": s.get("ecapa_mean"),
                "ecapa_min": s.get("ecapa_min"),
                "mcd_distance": s.get("mcd_distance"),
                "mel_distance": s.get("mel_distance"),
                "mfcc_distance": s.get("mfcc_distance"),
                "f0_correlation": s.get("f0_correlation"),
                "f0_rmse": s.get("f0_rmse"),
                "clipping_ratio": s.get("clipping_ratio"),
                "high_frequency_ratio": s.get("high_frequency_ratio"),
                "speech_ratio": s.get("speech_ratio"),
            },
            "warnings": s.get("warnings", []),
            "missing_metrics": s.get("missing_metrics", []),
            "absolute_score": s.get("absolute_score", {}),
            "status": s.get("status", "OK"),
        }
        models_data.append(entry)

    return {
        "timestamp": timestamp,
        "num_models": len(results),
        "weights": config.get("final_score", {}),
        "models": models_data,
    }


def _save_top5_audio(results: List[Dict[str, Any]], output_dir: Path, top_n: int) -> None:
    """Copy top-N model audio files to results/top5/ directory."""
    top5_dir = output_dir / "top5"
    top5_dir.mkdir(exist_ok=True)

    tech_sorted = sorted(results, key=lambda r: r["scores"].get("technical_rank", 999))
    for rank, r in enumerate(tech_sorted[:top_n], 1):
        src = Path(r["model_path"])
        if src.exists():
            dst = top5_dir / f"{rank:02d}_{r['model_name']}.wav"
            shutil.copy2(src, dst)
            logger.info("  Top %d: %s", rank, dst)


def _save_comparison_audio(
    results: List[Dict[str, Any]],
    config: Dict[str, Any],
    output_dir: Path,
) -> None:
    """Generate comparison.wav: concatenated original + top model snippets.

    Format: [original 3s] [silence 0.5s] [model1 3s] [silence 0.5s] [model2 3s] ...
    """
    import soundfile as sf

    sr = config.get("audio", {}).get("sample_rate", 16000)
    window_sec = config.get("window", {}).get("seconds", 3.0)
    silence_sec = 0.5
    top_n = config.get("output", {}).get("top_n", 5)

    tech_sorted = sorted(results, key=lambda r: r["scores"].get("technical_rank", 999))

    # Load original
    original_path = config.get("original_audio", "./original.wav")
    if not Path(original_path).exists():
        logger.warning("Original audio not found for comparison: %s", original_path)
        return

    orig_audio, _ = prepare_analysis_audio(original_path, sr, normalize=False)
    orig_snippet = orig_audio[:int(sr * window_sec)]

    silence = np.zeros(int(sr * silence_sec))

    segments = [orig_snippet, silence]

    for rank, r in enumerate(tech_sorted[:top_n], 1):
        model_path = Path(r["model_path"])
        if model_path.exists():
            model_audio, _ = prepare_analysis_audio(str(model_path), sr, normalize=False)
            model_snippet = model_audio[:int(sr * window_sec)]
            segments.append(model_snippet)
            if rank < min(top_n, len(tech_sorted)):
                segments.append(silence)

    comparison = np.concatenate(segments)
    output_path = output_dir / "comparison.wav"
    sf.write(str(output_path), comparison, sr)
    logger.info("Comparison audio saved: %s", output_path)


def main():
    parser = argparse.ArgumentParser(
        description="RVC Voice Model Evaluator — Automatic voice model quality assessment",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python evaluate.py
  python evaluate.py --original ./original.wav
  python evaluate.py --input ./models
  python evaluate.py --config ./config.yaml
  python evaluate.py --device cuda
  python evaluate.py --device cpu
  python evaluate.py --self-test
        """,
    )
    parser.add_argument("--original", default=None, help="Path to original reference WAV")
    parser.add_argument("--input", default=None, help="Directory containing model WAV files")
    parser.add_argument("--config", default="./config.yaml", help="Path to config YAML")
    parser.add_argument("--device", default=None, help="Device: 'cuda' or 'cpu' (overrides config)")
    parser.add_argument("--self-test", action="store_true", help="Run self-baseline test only")
    parser.add_argument("--compare", nargs=2, metavar=("A", "B"), default=None,
                        help="Compare two models after evaluation (e.g. --compare 24 25)")
    parser.add_argument("--trend-dir", default=None,
                        help="Run overfit trend analysis on a folder of evaluation results")
    args = parser.parse_args()

    # Load config
    config = load_config(args.config)

    # Override device
    if args.device:
        if "runtime" not in config:
            config["runtime"] = {}
        config["runtime"]["device"] = args.device

    # Set reproducibility
    repro = config.get("reproducibility", {})
    set_seed(repro.get("seed", 42), repro.get("deterministic", True))

    # Self-test mode
    if args.self_test:
        self_test(config)
        return

    # Trend analysis mode (standalone, no evaluation needed)
    if args.trend_dir:
        run_trend_analysis(args.trend_dir)
        return

    # Resolve paths
    original_path = args.original or config.get("original_audio", "./original.wav")
    models_dir = args.input or config.get("models_dir", "./models")
    results_dir = config.get("results_dir", "./results")

    # Run evaluation
    result_dir = evaluate_models(config, original_path, models_dir, results_dir)

    # A/B comparison (runs after evaluation, uses this run's results)
    if args.compare:
        run_comparison(result_dir, args.compare[0], args.compare[1])


if __name__ == "__main__":
    main()
